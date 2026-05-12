"""
create_sample.py
----------------
Script auxiliar para criar um arquivo Excel de exemplo com dados fictícios
de histórico de compras corporativas.

Execute: python create_sample.py
Saída: ../sample_data/historico_compras.xlsx
"""

import os
import pandas as pd
from datetime import datetime, timedelta
import random

# Dados fictícios de histórico de compras
dados = [
    # Produto,       Modelo,                    Último Preço, Data da Compra
    ("Mouse",        "Logitech M170",             59.90,  "2024-11-15"),
    ("Mouse",        "Logitech MX Master 3",      499.00, "2024-08-20"),
    ("Mouse",        "Microsoft Arc Mouse",        189.00, "2024-03-10"),
    ("Teclado",      "Logitech K120",              89.90,  "2024-10-22"),
    ("Teclado",      "Dell KB216",                 79.00,  "2024-05-14"),
    ("Teclado",      "Redragon K552",             169.00,  "2024-12-01"),
    ("Notebook",     "Dell Inspiron 15 3511",    3499.00, "2024-09-05"),
    ("Notebook",     "Lenovo IdeaPad 3",         2899.00, "2024-06-18"),
    ("Monitor",      "LG 24MK430H",               899.00, "2024-11-30"),
    ("Monitor",      "Samsung S24F350",            750.00, "2024-04-25"),
    ("Headset",      "JBL Quantum 100",            199.00, "2024-10-08"),
    ("Headset",      "Razer BlackShark V2 X",      279.90, "2024-07-12"),
    ("Webcam",       "Logitech C920",              499.00, "2024-12-10"),
    ("Webcam",       "Logitech C270",              219.00, "2024-02-20"),
    ("Impressora",   "HP DeskJet 2776",            499.99, "2024-11-18"),
    ("Impressora",   "Canon PIXMA G2160",          799.00, "2024-08-05"),
    ("Cadeira",      "Secretlab TITAN Evo",       4999.00, "2024-10-15"),
    ("Cadeira",      "ThunderX3 TC3",             1199.00, "2024-05-30"),
    ("HD Externo",   "Seagate Expansion 1TB",      289.90, "2024-12-05"),
    ("HD Externo",   "WD My Passport 2TB",         419.00, "2024-09-22"),
    ("Pendrive",     "SanDisk Ultra 64GB",          59.90, "2025-01-10"),
    ("Pendrive",     "Kingston DataTraveler 32GB",  29.90, "2024-11-20"),
]

def criar_excel_exemplo():
    """Cria o arquivo Excel de exemplo."""
    
    # Monta o DataFrame
    df = pd.DataFrame(dados, columns=[
        "Produto", "Modelo", "Último Preço", "Data da Compra"
    ])
    
    # Formata Data da Compra como datetime
    df["Data da Compra"] = pd.to_datetime(df["Data da Compra"])
    
    # Define o caminho de saída
    output_dir = os.path.join(os.path.dirname(__file__), "..", "sample_data")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "historico_compras.xlsx")
    
    # Salva como Excel com formatação básica
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Histórico de Compras", index=False)
        
        # Formata a planilha
        ws = writer.sheets["Histórico de Compras"]
        
        # Ajusta largura das colunas
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 18
        
        # Formata cabeçalho
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
    
    print(f"✅ Arquivo de exemplo criado: {output_path}")
    print(f"📊 Total de registros: {len(df)}")
    print(f"🗂️  Categorias: {df['Produto'].nunique()}")
    print(f"\nColunas do arquivo:")
    for col in df.columns:
        print(f"  - {col}")
    
    return output_path


if __name__ == "__main__":
    criar_excel_exemplo()

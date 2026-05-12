"""
exporter.py
-----------
Módulo para exportar os resultados de busca em formato Excel (.xlsx).
Gera um arquivo formatado com cabeçalhos estilizados e dados estruturados.
"""

import io
import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime


# Cores para estilização
COLOR_HEADER_BG = "1E40AF"   # Azul escuro
COLOR_HEADER_FG = "FFFFFF"   # Branco
COLOR_ROW_ALT   = "EFF6FF"   # Azul claríssimo (linhas alternadas)
COLOR_BORDER    = "BFDBFE"   # Azul claro para bordas


def _apply_border(cell):
    """Aplica borda fina em uma célula."""
    thin = Side(style="thin", color=COLOR_BORDER)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_excel(resultados: List[Dict[str, Any]]) -> bytes:
    """
    Gera um arquivo Excel (.xlsx) com os resultados de busca formatados.

    Args:
        resultados: Lista de dicts com campos do resultado de busca

    Returns:
        Bytes do arquivo Excel gerado
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados de Compras"

    # ── Título do relatório ─────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    titulo_cell = ws["A1"]
    titulo_cell.value = "📊 Relatório de Melhores Opções de Compra"
    titulo_cell.font = Font(bold=True, size=14, color=COLOR_HEADER_FG)
    titulo_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Subtítulo com data ──────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    data_cell = ws["A2"]
    data_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    data_cell.font = Font(italic=True, size=10, color="6B7280")
    data_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Linha em branco
    ws.row_dimensions[3].height = 8

    # ── Cabeçalhos da tabela ────────────────────────────────────────────────
    headers = [
        "Categoria",
        "Modelo Buscado",
        "Produto Encontrado",
        "Marketplace",
        "Preço (R$)",
        "Prazo de Entrega",
        "Link para Compra",
    ]
    header_row = 4
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG, size=11)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _apply_border(cell)
    ws.row_dimensions[header_row].height = 24

    # ── Dados ────────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
    normal_font = Font(size=10)
    link_font = Font(size=10, color="3B82F6", underline="single")

    for row_idx, item in enumerate(resultados, start=header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        row_fill = alt_fill if is_alt else PatternFill("none")

        values = [
            item.get("produto", ""),
            item.get("modelo_buscado", ""),
            item.get("titulo", ""),
            item.get("marketplace", ""),
            item.get("preco", 0),
            item.get("prazo", "Consultar"),
            item.get("link", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (3, 7) else "center",
                vertical="center",
                wrap_text=(col_idx == 3)
            )
            _apply_border(cell)

            # Estilo especial para preço
            if col_idx == 5:
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(size=10, bold=True, color="16A34A")
            # Link clicável
            elif col_idx == 7 and value and value != "#":
                cell.hyperlink = value
                cell.value = "Clique para ver"
                cell.font = link_font
            else:
                cell.font = normal_font

            if is_alt and col_idx != 5 and col_idx != 7:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 20

    # ── Larguras de coluna ─────────────────────────────────────────────────
    col_widths = [15, 20, 40, 15, 14, 20, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes (cabeçalho fixo) ──────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Auto-filter ──────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A{header_row}:G{header_row + len(resultados)}"

    # Salva em buffer de memória
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
